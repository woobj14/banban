# generator.py — S.Y. 담당: Excel 생성 엔진 (v2 — 합치기 지원)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import io
import math

# ─────────────────────────────────────────────────────────────────────────────
# 스타일 상수
# ─────────────────────────────────────────────────────────────────────────────

_THIN      = Side(border_style="thin", color="000000")
_ALL       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL  = PatternFill("solid", fgColor=Color(theme=4, tint=0.7999816888943144))
_AC        = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ACW       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# 본문 섹션(서론/본론/결론) 색상 — hdr: 헤더 배경, box: 요약 박스 배경
_SEC_COLORS = {
    "서론": {"hdr": "1E40AF", "box": "DBEAFE"},   # 파란 계열
    "본론": {"hdr": "166534", "box": "DCFCE7"},   # 초록 계열
    "결론": {"hdr": "92400E", "box": "FEF3C7"},   # 황토 계열
}
_SEC_DEFAULT = {"hdr": "374151", "box": "F3F4F6"}  # 알 수 없는 섹션

PAGE_WORD = 25
PAGE_DLG  = 25
PAGE_TEXT = 15


def _fk(size=12, bold=False):
    return Font(name="맑은 고딕", size=size, bold=bold)

def _fe(size=12, bold=True):
    return Font(name="Arial", size=size, bold=bold, color="1F1F1F")

def _hcell(cell, value):
    cell.value = value; cell.font = _fk(12, True)
    cell.fill  = _HDR_FILL; cell.alignment = _AC; cell.border = _ALL

_COLS = "ABCDEFGHIJ"

def _merged_header(ws, row: int, c1: int, c2: int, value: str):
    """c1~c2 열을 병합하고 헤더 스타일(중앙정렬+배경) 적용"""
    ws.merge_cells(f"{_COLS[c1-1]}{row}:{_COLS[c2-1]}{row}")
    cell = ws.cell(row=row, column=c1)
    cell.value     = value
    cell.font      = _fk(14, True)   # 구분 강조 14pt
    cell.fill      = _HDR_FILL
    cell.alignment = _ACW            # 중앙 + 자동줄바꿈
    cell.border    = _ALL

def _ncell(cell, value, size=10):
    cell.value = value; cell.font = _fk(size, True)
    cell.alignment = _AC; cell.border = _ALL

def _ecell(cell, value, h="center", size=10, bold=True):
    cell.value = value; cell.font = _fe(size, bold)
    cell.alignment = _AC if h == "center" else _AL; cell.border = _ALL

def _kcell(cell, value, h="center", size=10, bold=True, wrap=False):
    cell.value = value; cell.font = _fk(size, bold)
    cell.alignment = _ACW if wrap else (_AC if h == "center" else _AL)
    cell.border = _ALL

def _bcell(cell):
    cell.border = _ALL

def _page_setup(ws):
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9   # A4
    # 여백 (인치) — 위/아래 0.5", 왼/오른 0.3", 머리글/바닥글 0.5"
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.left   = 0.3
    ws.page_margins.right  = 0.3
    ws.page_margins.header = 0.5
    ws.page_margins.footer = 0.5

def _auto_row_h(en: str, kr: str,
                en_w: float = 36.0, kr_w: float = 34.0,
                font_size: int = 12, min_h: float = 20.0) -> float:
    """영한 텍스트 길이 기반 적정 행 높이 추정 (Excel pts).
    한글은 영문의 약 1.9배 너비로 계산."""
    line_pt = font_size * 1.5  # 12pt → 18pt per line (font + leading)

    def _nlines(text: str, col_w: float) -> int:
        if not text:
            return 1
        # 텍스트 내 한글 비율로 평균 문자 너비 보정
        kr_ratio     = sum(1 for c in text if '가' <= c <= '힣') / max(len(text), 1)
        char_w       = 1.0 + kr_ratio * 0.9     # 영문만=1.0, 한글만=1.9
        chars_per_ln = max(1, col_w / char_w)
        return max(1, math.ceil(len(text) / chars_per_ln))

    lines = max(_nlines(str(en or ""), en_w),
                _nlines(str(kr or ""), kr_w))
    return max(min_h, lines * line_pt + 4)

def _safe_name(name: str) -> str:
    """엑셀 시트명 31자 제한 + 금지 문자 제거"""
    for c in r"\/?*[]:'":
        name = name.replace(c, "")
    return name[:31]


# ─────────────────────────────────────────────────────────────────────────────
# 시트 채우기 함수 (워크시트 객체를 받음 → 재사용 가능)
# ─────────────────────────────────────────────────────────────────────────────

def _set_word_cols(ws):
    ws.column_dimensions["A"].width = 2.83
    ws.column_dimensions["B"].width = 39.5
    ws.column_dimensions["C"].width = 2.83
    ws.column_dimensions["D"].width = 2.83
    ws.column_dimensions["E"].width = 39.5


def fill_words(ws, title: str, words: list, label_c: str = "단어"):
    """단어 시트 채우기 (영어 + 한국어 반반)"""
    _set_word_cols(ws)
    row, first = 1, True
    if not words:
        _merged_header(ws, 1, 1, 2, title)
        _merged_header(ws, 1, 4, 5, title)
        _page_setup(ws); return

    for chunk_start in range(0, len(words), PAGE_WORD):
        chunk = words[chunk_start:chunk_start + PAGE_WORD]
        _merged_header(ws, row, 1, 2, title)
        _merged_header(ws, row, 4, 5, title)
        ws.row_dimensions[row].height = 28
        if first:
            c3 = ws.cell(row=row, column=3)
            c3.value = label_c; c3.font = _fk(14, True); c3.alignment = _AC
            first = False
        row += 1
        for i, (en, kr) in enumerate(chunk):
            _ncell(ws.cell(row=row, column=1), i + 1)
            _ecell(ws.cell(row=row, column=2), en)
            _ncell(ws.cell(row=row, column=4), i + 1)
            _kcell(ws.cell(row=row, column=5), kr, wrap=True)
            ws.row_dimensions[row].height = _auto_row_h(
                en, kr, en_w=39.5, kr_w=39.5, font_size=10, min_h=16)
            row += 1
    _page_setup(ws)


def fill_test_meaning(ws, title: str, words: list):
    """단어 테스트 뜻쓰기: 영어 보임 / 한국어 빈칸"""
    _set_word_cols(ws)
    row = 1
    if not words:
        _merged_header(ws, 1, 1, 2, title)
        _merged_header(ws, 1, 4, 5, title)
        _page_setup(ws); return
    for chunk_start in range(0, len(words), PAGE_WORD):
        chunk = words[chunk_start:chunk_start + PAGE_WORD]
        _merged_header(ws, row, 1, 2, title)
        _merged_header(ws, row, 4, 5, title)
        ws.row_dimensions[row].height = 28; row += 1
        for i, (en, _) in enumerate(chunk):
            _ncell(ws.cell(row=row, column=1), i + 1)
            _ecell(ws.cell(row=row, column=2), en)
            _ncell(ws.cell(row=row, column=4), i + 1)
            _bcell(ws.cell(row=row, column=5))
            ws.row_dimensions[row].height = _auto_row_h(
                en, "", en_w=39.5, kr_w=39.5, font_size=10, min_h=16)
            row += 1
    _page_setup(ws)


def fill_test_word(ws, title: str, words: list):
    """단어 테스트 단어쓰기: 한국어 보임 / 영어 빈칸"""
    _set_word_cols(ws)
    row = 1
    if not words:
        _merged_header(ws, 1, 1, 2, title)
        _merged_header(ws, 1, 4, 5, title)
        _page_setup(ws); return
    for chunk_start in range(0, len(words), PAGE_WORD):
        chunk = words[chunk_start:chunk_start + PAGE_WORD]
        _merged_header(ws, row, 1, 2, title)
        _merged_header(ws, row, 4, 5, title)
        ws.row_dimensions[row].height = 28; row += 1
        for i, (_, kr) in enumerate(chunk):
            _ncell(ws.cell(row=row, column=1), i + 1)
            _bcell(ws.cell(row=row, column=2))
            _ncell(ws.cell(row=row, column=4), i + 1)
            _kcell(ws.cell(row=row, column=5), kr, wrap=True)
            ws.row_dimensions[row].height = _auto_row_h(
                "", kr, en_w=39.5, kr_w=39.5, font_size=10, min_h=16)
            row += 1
    _page_setup(ws)


def fill_dialogues(ws, title: str, dialogues: list):
    """대화문 시트 채우기"""
    ws.column_dimensions["A"].width = 2.83
    ws.column_dimensions["B"].width = 39.5
    ws.column_dimensions["C"].width = 2.83
    ws.column_dimensions["D"].width = 2.83
    ws.column_dimensions["E"].width = 39.5

    flat = []
    for dlg in dialogues:
        flat.append(("title", dlg["title"], dlg["title"]))
        for en, kr in dlg.get("lines", []):
            flat.append(("line", en, kr))

    row, first = 1, True
    if not flat:
        _merged_header(ws, 1, 1, 2, title)
        _merged_header(ws, 1, 4, 5, title)
        _page_setup(ws); return

    for page_start in range(0, len(flat), PAGE_DLG):
        chunk = flat[page_start:page_start + PAGE_DLG]
        _merged_header(ws, row, 1, 2, title)
        _merged_header(ws, row, 4, 5, title)
        ws.row_dimensions[row].height = 44 if first else 28
        if first:
            c3 = ws.cell(row=row, column=3)
            c3.value = "대화문"; c3.font = _fk(14, True)
            c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            first = False
        row += 1
        for i, (rtype, en, kr) in enumerate(chunk):
            _ncell(ws.cell(row=row, column=1), i + 1, size=10)
            _ncell(ws.cell(row=row, column=4), i + 1, size=10)
            if rtype == "title":
                for col, text in [(2, en), (5, kr)]:
                    c = ws.cell(row=row, column=col)
                    c.value = text; c.font = _fk(14, True)   # 강조: 14pt
                    c.alignment = _ACW; c.border = _ALL; c.fill = _HDR_FILL
                ws.row_dimensions[row].height = _auto_row_h(
                    en, kr, en_w=39.5, kr_w=39.5, font_size=14, min_h=22)
            else:
                for col, text in [(2, en), (5, kr)]:
                    c = ws.cell(row=row, column=col)
                    c.value = text; c.font = _fk(10)          # 내용: 10pt
                    c.alignment = _AL; c.border = _ALL
                ws.row_dimensions[row].height = _auto_row_h(
                    en, kr, en_w=39.5, kr_w=39.5, font_size=10, min_h=16)
            row += 1
    _page_setup(ws)


def fill_text(ws, title: str, text_data: dict):
    """본문 시트 채우기.

    text_data에 'sections' 키가 있으면 서론/본론/결론 분할 레이아웃 적용.
    없으면 기존 페이지 분할 레이아웃(하위 호환).
    """
    ws.column_dimensions["A"].width = 2.83
    ws.column_dimensions["B"].width = 39.5
    ws.column_dimensions["C"].width = 2.83
    ws.column_dimensions["D"].width = 2.83
    ws.column_dimensions["E"].width = 39.5

    sentences = text_data.get("sentences", [])
    sections  = text_data.get("sections")     # None 또는 list
    title_en  = text_data.get("title_en", "")
    title_kr  = text_data.get("title_kr", "")

    row = 1

    if not sentences:
        _merged_header(ws, 1, 1, 2, title)
        _merged_header(ws, 1, 4, 5, title)
        _page_setup(ws); return

    if sections:
        # ── 섹션(서론/본론/결론) 레이아웃 ─────────────────────
        _merged_header(ws, row, 1, 2, title)
        _merged_header(ws, row, 4, 5, title)
        c3 = ws.cell(row=row, column=3)
        c3.value = "본문"; c3.font = _fk(14, True); c3.alignment = _AC
        ws.row_dimensions[row].height = 28
        row += 1

        # 본문 제목 행
        for col, text in [(2, title_en), (5, title_kr)]:
            c = ws.cell(row=row, column=col)
            c.value = text; c.font = _fk(14, True)             # 강조: 14pt
            c.alignment = _ACW; c.border = _ALL; c.fill = _HDR_FILL
        ws.row_dimensions[row].height = _auto_row_h(
            title_en, title_kr, en_w=39.5, kr_w=39.5, font_size=14, min_h=22)
        row += 1

        sent_idx = 0
        for sec in sections:
            label     = sec.get("label", "")
            colors    = _SEC_COLORS.get(label, _SEC_DEFAULT)
            sec_sents = sec.get("sentences", [])

            # 섹션 헤더 행 (A~E 병합, 색상 배경, 흰 글씨) — 강조 14pt
            ws.merge_cells(f"A{row}:E{row}")
            c = ws.cell(row=row, column=1)
            c.value     = f"◆ {label}"
            c.font      = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor=colors["hdr"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _ALL
            ws.row_dimensions[row].height = 26
            row += 1

            # 문장 행
            for en, kr in sec_sents:
                sent_idx += 1
                _ncell(ws.cell(row=row, column=1), sent_idx)
                _ncell(ws.cell(row=row, column=4), sent_idx)
                for col, text in [(2, en), (5, kr)]:
                    c = ws.cell(row=row, column=col)
                    c.value = text; c.font = _fk(10)            # 내용: 10pt
                    c.alignment = _AL; c.border = _ALL
                ws.row_dimensions[row].height = _auto_row_h(
                    en, kr, en_w=39.5, kr_w=39.5, font_size=10, min_h=28)
                row += 1

            # 요약 박스 레이블 행 (A~E 병합)
            ws.merge_cells(f"A{row}:E{row}")
            c = ws.cell(row=row, column=1)
            c.value     = f"✏  {label} 핵심 내용 정리"
            c.font      = Font(name="맑은 고딕", size=10, bold=True, color="374151")
            c.fill      = PatternFill("solid", fgColor=colors["box"])
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1)
            c.border    = _ALL
            ws.row_dimensions[row].height = 20
            row += 1

            # 빈 쓰기 행 × 2 (같은 배경)
            for _ in range(2):
                ws.merge_cells(f"A{row}:E{row}")
                c = ws.cell(row=row, column=1)
                c.fill   = PatternFill("solid", fgColor=colors["box"])
                c.border = _ALL
                ws.row_dimensions[row].height = 40
                row += 1

            # 간격 행
            ws.row_dimensions[row].height = 8
            row += 1

    else:
        # ── 기존 페이지 분할 레이아웃 (하위 호환) ─────────────
        first = True
        for page_start in range(0, len(sentences), PAGE_TEXT):
            chunk = sentences[page_start:page_start + PAGE_TEXT]
            _merged_header(ws, row, 1, 2, title)
            _merged_header(ws, row, 4, 5, title)
            ws.row_dimensions[row].height = 28
            if first:
                c3 = ws.cell(row=row, column=3)
                c3.value = "본문"; c3.font = _fk(14, True); c3.alignment = _AC
                first = False; row += 1
                for col, text in [(2, title_en), (5, title_kr)]:
                    c = ws.cell(row=row, column=col)
                    c.value = text; c.font = _fk(14, True)     # 강조: 14pt
                    c.alignment = _ACW; c.border = _ALL; c.fill = _HDR_FILL
                ws.row_dimensions[row].height = _auto_row_h(
                    title_en, title_kr, en_w=39.5, kr_w=39.5, font_size=14, min_h=22)
            row += 1
            for i, (en, kr) in enumerate(chunk):
                num = page_start + i + 1
                _ncell(ws.cell(row=row, column=1), num)
                _ncell(ws.cell(row=row, column=4), num)
                for col, text in [(2, en), (5, kr)]:
                    c = ws.cell(row=row, column=col)
                    c.value = text; c.font = _fk(10)            # 내용: 10pt
                    c.alignment = _AL; c.border = _ALL
                ws.row_dimensions[row].height = _auto_row_h(
                    en, kr, en_w=39.5, kr_w=39.5, font_size=10, min_h=28)
                row += 1

    _page_setup(ws)


# ─────────────────────────────────────────────────────────────────────────────
# 표지
# ─────────────────────────────────────────────────────────────────────────────

def _fill_cover(ws, meta):
    """표지 내용 채우기 — 오른쪽 절반 배치 / A4 반절 책자 대응.

    레이아웃:
      A~C열 : 좌측 여백 (접으면 뒤표지 / 뒷면)
      D열   : 표지 텍스트 (우측 정렬)

    색상:
      진한 검정 (1F2937) → '스스로 발화 노트' / 출판사 / 저자
      초록 강조 (2D8A45) → 섹션 유형(단어·대화문·본문) / 학년 / 과
    """
    grade, pub, author, chapter = (
        meta["grade"], meta["publisher"], meta["author"], meta["chapter"]
    )
    section = meta.get("section", "본문")

    _DARK   = "1F2937"   # 거의 검정
    _ACCENT = "2D8A45"   # 초록 강조

    # ── 열 너비 ────────────────────────────────────────────
    # 총 너비 ≈ 77 units (본문/단어 시트와 동일, A4 가로 가득 채움)
    # A~C: 좌측 절반 스페이서(접으면 뒤표지/뒷면) = 39 units
    # D  : 우측 절반 표지 내용(접으면 앞표지)     = 35 units
    # E  : 우측 여백                               =  3 units
    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 13.0   # 13 × 3 = 39
    ws.column_dimensions["D"].width = 35.0
    ws.column_dimensions["E"].width = 3.0

    # ── 행 높이 ────────────────────────────────────────────
    for r in range(1, 17):
        ws.row_dimensions[r].height = 30

    _AR = Alignment(horizontal="right", vertical="center")

    # ── 텍스트 배치 (행, 내용, 폰트크기, bold, 색상) ────────
    layout = [
        (3,  "스스로",        34, True, _DARK),    # ─┐
        (4,  "발화",          34, True, _DARK),    #  │ 타이틀 (검정)
        (5,  "노트",          34, True, _DARK),    # ─┘
        (7,  section,        28, True, _ACCENT),  # 단어·대화문·본문 (강조)
        (8,  grade,          24, True, _ACCENT),  # 중1·중2… (강조)
        (9,  f"{chapter}과", 24, True, _ACCENT),  # 4과… (강조)
        (11, pub,            20, True, _DARK),    # ─┐
        (12, author,         20, True, _DARK),    # ─┘ 출판사·저자 (검정)
    ]

    for row, text, size, bold, color in layout:
        c            = ws.cell(row=row, column=4)
        c.value      = text
        c.font       = Font(name="맑은 고딕", size=size, bold=bold, color=color)
        c.alignment  = _AR

    # A4 기준 인쇄 영역: 전체 너비 포함 (접으면 우측=표지)
    ws.print_area = "A1:E16"
    _page_setup(ws)
    # 표지: 가로 1페이지에 맞게 자동 스케일 (세로는 자동)
    ws.sheet_properties = WorksheetProperties(
        pageSetUpPr=PageSetupProperties(fitToPage=True)
    )
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0


# ─────────────────────────────────────────────────────────────────────────────
# Public API 1: 단일 노트 생성 (기존 기능 유지)
# ─────────────────────────────────────────────────────────────────────────────

def generate(meta: dict, words: list, dialogues: list, text_data: dict) -> bytes:
    """섹션별 표지 포함 단일 스발노트 엑셀 생성.

    시트 순서 예시 (전체 유형):
      단어 표지 → 단어 → 단어 테스트 뜻쓰기 → 단어 테스트 단어쓰기
      → 대화문 표지 → 대화문
      → 본문 표지 → 본문
    """
    wb    = Workbook()
    grade, pub, author, chapter = meta["grade"], meta["publisher"], meta["author"], meta["chapter"]
    base  = f"{grade} {pub} {author} {chapter}과"

    has_w = bool(words)
    has_d = bool(dialogues)
    has_t = bool(text_data.get("sentences"))
    multi = sum([has_w, has_d, has_t]) > 1   # 2개 이상이면 표지 이름에 유형 붙임

    first = True   # True이면 wb.active(기본 Sheet) 재활용

    def _add_cover(section_type: str):
        nonlocal first
        cover_name = f"{section_type} 표지" if multi else "표지"
        if first:
            ws      = wb.active
            ws.title = cover_name
            first   = False
        else:
            ws = wb.create_sheet(cover_name)
        _fill_cover(ws, {**meta, "section": section_type})

    if has_w:
        _add_cover("단어")
        ws = wb.create_sheet("단어");                fill_words(ws, f"{base} 단어", words)
        ws = wb.create_sheet("단어 테스트 뜻쓰기");  fill_test_meaning(ws, f"{base} 단어", words)
        ws = wb.create_sheet("단어 테스트 단어쓰기"); fill_test_word(ws, f"{base} 단어", words)

    if has_d:
        _add_cover("대화문")
        ws = wb.create_sheet("대화문");              fill_dialogues(ws, f"{base} 대화문", dialogues)

    if has_t:
        _add_cover("본문")
        ws = wb.create_sheet("본문");                fill_text(ws, f"{base} 본문", text_data)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Public API 2: 다중 노트 합치기 (신규)
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_plan(note: dict) -> list[dict]:
    """노트 하나가 생성할 시트 목록 반환 (표지 포함).

    반환 형식:
      [{"icon": str, "name": str, "fn": callable, "args": tuple}, ...]

    표지 시트는 _fill_cover(ws, meta_dict) 형태로 호출됨.
    """
    grade, pub, author = note["grade"], note["publisher"], note["author"]
    chapter = note["chapter"]
    ctype   = note["content_type"]
    base    = f"{grade} {pub} {author} {chapter}과"
    ch      = f"{chapter}과" if chapter else note.get("title", "노트")[:6]

    words     = note.get("words", [])
    dialogues = note.get("dialogues", [])
    text_data = note.get("text_data", {})

    cover_meta_base = {
        "grade": grade, "publisher": pub,
        "author": author, "chapter": chapter,
    }

    # 이 노트에 몇 개의 섹션이 있는지 파악
    sections = []
    if ctype in ("단어",   "전체") and words:                   sections.append("단어")
    if ctype in ("대화문", "전체") and dialogues:               sections.append("대화문")
    if ctype in ("본문",   "전체") and text_data.get("sentences"): sections.append("본문")
    multi = len(sections) > 1

    plan = []

    if "단어" in sections:
        cover_name = _safe_name(f"{ch}_단어표지" if multi else f"{ch}_표지")
        plan.append({"icon": "📋", "name": cover_name,
                     "fn": _fill_cover,
                     "args": ({**cover_meta_base, "section": "단어"},)})
        plan.append({"icon": "📝", "name": _safe_name(f"{ch}_단어"),
                     "fn": fill_words,        "args": (f"{base} 단어", words)})
        plan.append({"icon": "✏️", "name": _safe_name(f"{ch}_뜻쓰기"),
                     "fn": fill_test_meaning, "args": (f"{base} 단어", words)})
        plan.append({"icon": "✏️", "name": _safe_name(f"{ch}_단어쓰기"),
                     "fn": fill_test_word,    "args": (f"{base} 단어", words)})

    if "대화문" in sections:
        cover_name = _safe_name(f"{ch}_대화문표지" if multi else f"{ch}_표지")
        plan.append({"icon": "📋", "name": cover_name,
                     "fn": _fill_cover,
                     "args": ({**cover_meta_base, "section": "대화문"},)})
        plan.append({"icon": "💬", "name": _safe_name(f"{ch}_대화문"),
                     "fn": fill_dialogues,    "args": (f"{base} 대화문", dialogues)})

    if "본문" in sections:
        cover_name = _safe_name(f"{ch}_본문표지" if multi else f"{ch}_표지")
        plan.append({"icon": "📋", "name": cover_name,
                     "fn": _fill_cover,
                     "args": ({**cover_meta_base, "section": "본문"},)})
        plan.append({"icon": "📖", "name": _safe_name(f"{ch}_본문"),
                     "fn": fill_text,         "args": (f"{base} 본문", text_data)})

    return plan


def sheet_preview(notes: list) -> list[dict]:
    """합치기 전 시트 미리보기 목록 반환 (UI 표시용)."""
    result = []
    seen   = {}
    for note in notes:
        for item in _sheet_plan(note):
            name = item["name"]
            # 이름 충돌 처리
            if name in seen:
                seen[name] += 1
                name = _safe_name(f"{name}_{seen[name]}")
            else:
                seen[name] = 1
            result.append({"icon": item["icon"], "name": name,
                           "note_title": note.get("title", "")})
    return result


def generate_combined(notes: list) -> bytes:
    """
    여러 노트를 하나의 엑셀 파일로 합침.
    notes: library.get_note() 결과 리스트 (순서 = 시트 순서)
    """
    wb   = Workbook()
    seen = {}

    for note in notes:
        for item in _sheet_plan(note):
            # 시트 이름 중복 방지
            name = item["name"]
            if name in seen:
                seen[name] += 1
                name = _safe_name(f"{name}_{seen[name]}")
            else:
                seen[name] = 1

            ws = wb.create_sheet(name)
            item["fn"](ws, *item["args"])

    # 빈 기본 시트 제거
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()
